import os
import io
import json
from typing import TYPE_CHECKING, Literal

import pandas as pd

import openpyxl
from openpyxl import styles as openpyxl_styles
from openpyxl.utils import get_column_letter

from flask import Blueprint, render_template, request, abort, flash, url_for, current_app, Response
from flask_htmx import make_response
from flask_login import login_required

from opengsync_db import models, PAGE_LIMIT, db_session
from opengsync_db.categories import HTTPResponse, LabProtocol, PoolStatus, LibraryStatus, PrepStatus, SeqRequestStatus

from .... import db, forms, logger, htmx_route  # noqa
from ....tools.spread_sheet_components import TextColumn


if TYPE_CHECKING:
    current_user: models.User = None    # type: ignore
else:
    from flask_login import current_user

lab_preps_htmx = Blueprint("lab_preps_htmx", __name__, url_prefix="/api/hmtx/lab_preps/")


@lab_preps_htmx.route("get", methods=["GET"], defaults={"page": 0})
@lab_preps_htmx.route("get/<int:page>", methods=["GET"])
@db_session(db)
@login_required
def get(page: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    sort_by = request.args.get("sort_by", "id")
    sort_order = request.args.get("sort_order", "desc")
    descending = sort_order == "desc"
    offset = PAGE_LIMIT * page

    if (status_in := request.args.get("status_id_in")) is not None:
        status_in = json.loads(status_in)
        try:
            status_in = [PrepStatus.get(int(status)) for status in status_in]
        except ValueError:
            return abort(HTTPResponse.BAD_REQUEST.id)
    
        if len(status_in) == 0:
            status_in = None

    if (protocol_in := request.args.get("protocol_id_in")) is not None:
        protocol_in = json.loads(protocol_in)
        try:
            protocol_in = [LabProtocol.get(int(protocol_)) for protocol_ in protocol_in]
        except ValueError:
            return abort(HTTPResponse.BAD_REQUEST.id)
    
        if len(protocol_in) == 0:
            protocol_in = None

    lab_preps, n_pages = db.get_lab_preps(
        status_in=status_in, protocol_in=protocol_in,
        offset=offset, limit=PAGE_LIMIT, sort_by=sort_by, descending=descending, count_pages=True
    )
    
    return render_template(
        "components/tables/lab_prep.html", lab_preps=lab_preps, n_pages=n_pages, active_page=page,
        sort_by=sort_by, sort_order=sort_order, status_in=status_in, protocol_in=protocol_in,
    )


@htmx_route(lab_preps_htmx, db=db)
def table_query():
    if (word := request.args.get("name")) is not None:
        field_name = "name"
    elif (word := request.args.get("id")) is not None:
        field_name = "id"
    elif (word := request.args.get("creator_id")) is not None:
        field_name = "creator_id"
    else:
        return abort(HTTPResponse.BAD_REQUEST.id)

    if (status_in := request.args.get("status_id_in")) is not None:
        status_in = json.loads(status_in)
        try:
            status_in = [PrepStatus.get(int(status)) for status in status_in]
        except ValueError:
            return abort(HTTPResponse.BAD_REQUEST.id)
    
        if len(status_in) == 0:
            status_in = None

    if (protocol_in := request.args.get("protocol_id_in")) is not None:
        protocol_in = json.loads(protocol_in)
        try:
            protocol_in = [LabProtocol.get(int(protocol_)) for protocol_ in protocol_in]
        except ValueError:
            return abort(HTTPResponse.BAD_REQUEST.id)
    
        if len(protocol_in) == 0:
            protocol_in = None

    lab_preps: list[models.LabPrep] = []
    if field_name == "name":
        lab_preps = db.query_lab_preps(name=word, protocol_in=protocol_in, status_in=status_in)
    elif field_name == "id":
        try:
            _id = int(word)
            if (lab_prep := db.get_lab_prep(_id)) is not None:
                lab_preps.append(lab_prep)
        except ValueError:
            pass
    elif field_name == "creator_id":
        lab_preps = db.query_lab_preps(creator=word, protocol_in=protocol_in, status_in=status_in)

    return make_response(
        render_template(
            "components/tables/lab_prep.html",
            current_query=word, active_query_field=field_name,
            lab_preps=lab_preps, protocol_in=protocol_in, status_in=status_in
        )
    )


@htmx_route(lab_preps_htmx, db=db)
def get_form(form_type: Literal["create", "edit"]):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if form_type not in ["create", "edit"]:
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    if (lab_prep_id := request.args.get("lab_prep_id")) is not None:
        try:
            lab_prep_id = int(lab_prep_id)
        except ValueError:
            return abort(HTTPResponse.BAD_REQUEST.id)
        
        if form_type != "edit":
            return abort(HTTPResponse.BAD_REQUEST.id)
        
        if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
            return abort(HTTPResponse.NOT_FOUND.id)
        
        return forms.models.LabPrepForm(form_type=form_type, lab_prep=lab_prep).make_response()
    
    # seq_request_id must be provided if form_type is "edit"
    if form_type == "edit":
        return abort(HTTPResponse.BAD_REQUEST.id)

    return forms.models.LabPrepForm(form_type=form_type).make_response()


@htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def create():
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    form = forms.models.LabPrepForm(formdata=request.form, form_type="create")
    return form.process_request(current_user)


@htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def edit(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    form = forms.models.LabPrepForm(formdata=request.form, form_type="edit", lab_prep=lab_prep)
    return form.process_request(current_user)


@htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def complete(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    for pool in lab_prep.pools:
        pool.status = PoolStatus.STORED
        pool = db.update_pool(pool)

    for library in lab_prep.libraries:
        is_prepared = True
        for sr_library in library.seq_request.libraries:
            is_prepared = sr_library.status >= LibraryStatus.POOLED and is_prepared
            if not is_prepared:
                break
        if is_prepared:
            library.seq_request.status = SeqRequestStatus.PREPARED
            library = db.update_library(library)

    lab_prep.status = PrepStatus.COMPLETED
    lab_prep = db.update_lab_prep(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if lab_prep.status != PrepStatus.PREPARING:
        flash("Cannot delete completed prep.", "warning")
        return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))
    
    db.delete_lab_prep(lab_prep_id=lab_prep.id)
    flash("Lab prep deleted!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_preps"))


@htmx_route(lab_preps_htmx, db=db, methods=["POST"])
def uncomplete(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    for library in lab_prep.libraries:
        if library.status == LibraryStatus.POOLED:
            library.status = LibraryStatus.PREPARING

    lab_prep.status = PrepStatus.PREPARING
    lab_prep = db.update_lab_prep(lab_prep)

    flash("Lab prep completed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def remove_library(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if lab_prep.status != PrepStatus.PREPARING:
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    if (library_id := request.args.get("library_id")) is None:
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    try:
        library_id = int(library_id)
    except ValueError:
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    db.remove_library_from_prep(lab_prep_id=lab_prep.id, library_id=library_id)

    flash("Library removed!", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@lab_preps_htmx.route("<int:lab_prep_id>/get_libraries", methods=["GET"], defaults={"page": 0})
@lab_preps_htmx.route("<int:lab_prep_id>/get_libraries/<int:page>", methods=["GET"])
@db_session(db)
@login_required
def get_libraries(lab_prep_id: int, page: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    sort_by = request.args.get("sort_by", "id")
    sort_order = request.args.get("sort_order", "desc")
    descending = sort_order == "desc"
    offset = PAGE_LIMIT * page

    libraries, n_pages = db.get_libraries(offset=offset, lab_prep_id=lab_prep_id, sort_by=sort_by, descending=descending, count_pages=True)
    
    return make_response(
        render_template(
            "components/tables/lab_prep-library.html",
            libraries=libraries, n_pages=n_pages, active_page=page,
            sort_by=sort_by, sort_order=sort_order, lab_prep=lab_prep
        )
    )


@htmx_route(lab_preps_htmx, db=db, methods=["GET"])
def download_template(lab_prep_id: int, direction: Literal["rows", "columns"]) -> Response:
    if direction not in ("rows", "columns"):
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if current_app.static_folder is None:
        logger.error("Static folder not set")
        raise ValueError("Static folder not set")
    
    filepath = os.path.join(current_app.static_folder, "resources", "templates", "library_prep", lab_prep.protocol.prep_file_name)

    if not os.path.exists(filepath):
        logger.error(f"File not found: {filepath}")
        return abort(HTTPResponse.NOT_FOUND.id)

    template = openpyxl.load_workbook(filepath)

    n = 12 if direction == "rows" else 8
    pattern = [True] * n + [False] * n

    def if_color(i):
        return pattern[i]

    active_sheet = template["prep_table"]
    column_mapping: dict[str, str] = {}
    
    for col_i in range(0, min(active_sheet.max_column, 96)):
        col = get_column_letter(col_i + 1)
        column_name = active_sheet[f"{col}1"].value
        column_mapping[column_name] = col
        
        for row_idx, cell in enumerate(active_sheet[col][1:]):
            if if_color(row_idx % (n * 2)):
                cell.fill = openpyxl_styles.PatternFill(start_color="ced4da", end_color="ced4da", fill_type="solid")
            else:
                cell.fill = openpyxl_styles.PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

    for row_idx, cell in enumerate(active_sheet[column_mapping["plate_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")

    for row_idx, cell in enumerate(active_sheet[column_mapping["index_well"]][1:]):
        if row_idx > 95:
            break
        cell.value = models.Plate.well_identifier(row_idx, num_cols=12, num_rows=8, flipped=direction == "columns")
        
    for i, library in enumerate(lab_prep.libraries):
        library_id_cell = active_sheet[f"{column_mapping['library_id']}{i + 2}"]
        library_name_cell = active_sheet[f"{column_mapping['library_name']}{i + 2}"]
        requestor_cell = active_sheet[f"{column_mapping['requestor']}{i + 2}"]
        sequence_i7_cell = active_sheet[f"{column_mapping['sequence_i7']}{i + 2}"]
        sequence_i5_cell = active_sheet[f"{column_mapping['sequence_i5']}{i + 2}"]
        # kit_i7_cell = active_sheet[f"{column_mapping['kit_i7']}{i + 2}"]
        # kit_i5_cell = active_sheet[f"{column_mapping['kit_i5']}{i + 2}"]
        name_i7_cell = active_sheet[f"{column_mapping['name_i7']}{i + 2}"]
        name_i5_cell = active_sheet[f"{column_mapping['name_i5']}{i + 2}"]
        library_id_cell.value = library.id
        library_name_cell.value = library.name
        requestor_cell.value = library.seq_request.requestor.name
        if len(library.indices) > 0:
            name_i7_cell.value = library.indices[0].name_i7
            name_i5_cell.value = library.indices[0].name_i5
        sequence_i7_cell.value = library.sequences_i7_str(";")
        sequence_i5_cell.value = library.sequences_i5_str(";")
        
    bytes_io = io.BytesIO()
    template.save(bytes_io)

    bytes_io.seek(0)
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
    return Response(
        bytes_io, mimetype=mimetype,
        headers={"Content-disposition": f"attachment; filename={lab_prep.name}_{lab_prep.protocol.abbreviation}_{direction}.xlsx"}
    )


@htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def prep_table_upload_form(lab_prep_id: int) -> Response:
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if request.method == "GET":
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep)
        return form.make_response()
    else:
        form = forms.workflows.library_prep.LibraryPrepForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(user=current_user)
    

@lab_preps_htmx.route("<int:lab_prep_id>/get_pools/<int:page>", methods=["GET"])
@lab_preps_htmx.route("<int:lab_prep_id>/get_pools", methods=["GET"], defaults={"page": 0})
@db_session(db)
@login_required
def get_pools(lab_prep_id: int, page: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    sort_by = request.args.get("sort_by", "id")
    sort_order = request.args.get("sort_order", "desc")
    descending = sort_order == "desc"
    offset = PAGE_LIMIT * page

    pools, n_pages = db.get_pools(
        lab_prep_id=lab_prep_id, offset=offset, sort_by=sort_by, descending=descending, count_pages=True
    )

    return make_response(
        render_template(
            "components/tables/lab_prep-pool.html",
            pools=pools, n_pages=n_pages, active_page=page,
            sort_by=sort_by, sort_order=sort_order, lab_prep=lab_prep
        )
    )


@htmx_route(lab_preps_htmx, db=db)
def get_files(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)

    return make_response(
        render_template(
            "components/file-list.html",
            files=lab_prep.files, delete="lab_preps_htmx.delete_file",
            delete_context={"lab_prep_id": lab_prep_id}
        )
    )


@htmx_route(lab_preps_htmx, db=db, methods=["DELETE"])
def delete_file(lab_prep_id: int, file_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if (file := db.get_file(file_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if file not in lab_prep.files:
        return abort(HTTPResponse.BAD_REQUEST.id)
    
    file_path = os.path.join(current_app.config["MEDIA_FOLDER"], file.path)
    if os.path.exists(file_path):
        os.remove(file_path)
    db.delete_file(file_id=file.id)

    logger.info(f"Deleted file '{file.name}' from prep (id='{lab_prep_id}')")
    flash(f"Deleted file '{file.name}' from prep.", "success")
    return make_response(redirect=url_for("lab_preps_page.lab_prep", lab_prep_id=lab_prep_id))


@htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def file_form(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if request.method == "GET":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.file.LabPrepAttachmentForm(lab_prep=lab_prep, formdata=request.form | request.files)
        return form.process_request(current_user)
    else:
        return abort(HTTPResponse.METHOD_NOT_ALLOWED.id)
    

@htmx_route(lab_preps_htmx, db=db)
def plates_tab(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    return make_response(render_template("components/plates.html", plates=lab_prep.plates, from_page=f"lab_prep@{lab_prep.id}"))
    

@htmx_route(lab_preps_htmx, db=db, methods=["GET", "POST"])
def comment_form(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if request.method == "GET":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep)
        return form.make_response()
    elif request.method == "POST":
        form = forms.comment.LabPrepCommentForm(lab_prep=lab_prep, formdata=request.form)
        return form.process_request(current_user)
    else:
        return abort(HTTPResponse.METHOD_NOT_ALLOWED.id)


@htmx_route(lab_preps_htmx, db=db)
def get_comments(lab_prep_id: int):
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)

    return make_response(
        render_template(
            "components/comment-list.html",
            comments=lab_prep.comments
        )
    )


@htmx_route(lab_preps_htmx, db=db)
def get_sample_pooling_table(lab_prep_id: int):
    if not current_user.is_insider():
        return abort(HTTPResponse.FORBIDDEN.id)
    
    if (lab_prep := db.get_lab_prep(lab_prep_id)) is None:
        return abort(HTTPResponse.NOT_FOUND.id)
    
    df = db.get_lab_prep_samples_df(lab_prep.id)

    mux_table = {
        "sample_name": [],
        "library_name": [],
        "sample_pool": [],
        "barcode": [],
        "pattern": [],
        "read": []
    }

    for _, row in df.iterrows():
        if row["mux_type_id"] is None:
            continue
        
        mux_table["sample_name"].append(row["sample_name"])
        mux_table["library_name"].append(row["library_name"])
        mux_table["sample_pool"].append(row["sample_pool"])
        if (mux := row.get("mux")) is None:
            mux_table["barcode"].append(None)
            mux_table["pattern"].append(None)
            mux_table["read"].append(None)
        else:
            mux_table["barcode"].append(mux.get("barcode"))
            mux_table["pattern"].append(mux.get("pattern"))
            mux_table["read"].append(mux.get("read"))

    df = pd.DataFrame(mux_table)
    if df["pattern"].isna().all():
        df = df.drop(columns=["pattern"])
    if df["read"].isna().all():
        df = df.drop(columns=["read"])
        
    columns = []
    for i, col in enumerate(df.columns):
        columns.append(
            TextColumn(
                col,
                col.replace("_", " ").title().replace("Id", "ID").replace("Cmo", "CMO"),
                {
                    "sample_name": 200,
                    "library_name": 250,
                    "sample_pool": 200,
                    "barcode": 100,
                    "read": 80,
                    "pattern": 150
                }.get(col, 100),
                max_length=1000
            )
        )

    return make_response(
        render_template(
            "components/itable.html", columns=columns,
            spreadsheet_data=df.replace(pd.NA, "").values.tolist(),
        )
    )
